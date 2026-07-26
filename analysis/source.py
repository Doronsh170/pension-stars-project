#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Shared access to the two source workbooks in the repository root.

`gemel-2010-2025.xlsx`  — the consolidated gemel-net extract: one row per
    fund x reporting month, 2010-2025, already restricted by the regulator's
    TARGET_POPULATION field to funds open to the general public. It carries
    the regulator's own SPECIALIZATION / SUB_SPECIALIZATION track labels.

`fund-risk-map.xlsx`    — the curated mapping table: FUND_NAME ->
    SPECIALIZATION, SUB_SPECIALIZATION and, most importantly, a single stable
    `רמת סיכון` (risk level) per fund. This is the column the study groups on.

Both files are read here so every stage sees the same universe and the same
exclusion rules.
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

DATA_XLSX = os.path.join(ROOT, "gemel-2010-2025.xlsx")
RISKMAP_XLSX = os.path.join(ROOT, "fund-risk-map.xlsx")

# Risk levels, least to most risky. This is the study's comparison dimension:
# funds are ranked against other funds of the same risk level, not against the
# whole market, so a "winner" is never just the fund that took more equity.
RISK_ORDER = ["נמוך מאוד", "נמוך", "נמוך-בינוני", "בינוני", "גבוה"]

# Fund families a saver actively chooses between. `מרכזית לפיצויים` (an employer
# severance reserve) and `מטרה אחרת` are excluded: the saver does not pick them.
FAMILIES = [
    "קרנות השתלמות",
    "תגמולים ואישית לפיצויים",
    "קופת גמל להשקעה",
    "קופת גמל להשקעה - חסכון לילד",
]

# Shorter, saver-facing names for the same families.
FAMILY_LABEL = {
    "קרנות השתלמות": "קרן השתלמות",
    "תגמולים ואישית לפיצויים": "קופת גמל",
    "קופת גמל להשקעה": "גמל להשקעה",
    "קופת גמל להשקעה - חסכון לילד": "חיסכון לכל ילד",
}

BLANK = "(ריק)"   # how the mapping table spells an empty cell


def norm(s):
    """Collapse whitespace; treat the mapping table's `(ריק)` as empty."""
    if s is None:
        return ""
    s = " ".join(str(s).split()).strip()
    return "" if s == BLANK else s


def norm_risk(s):
    """Normalise the risk label (the source has one `נמוך -בינוני` typo)."""
    s = norm(s)
    return s.replace("נמוך -בינוני", "נמוך-בינוני")


def excluded_fund(name):
    """Products that are not a savings choice a normal saver can make.

    * `בניהול אישי` / IRA — legal shells for member-directed portfolios. They
      report placeholder yields with no pooled fund return behind them.
    * `דמי מחלה` — employer sick-pay reserves, not a personal savings product.
    """
    return bool(name) and ("בניהול אישי" in name or "IRA" in name or "דמי מחלה" in name)


def _data_sheet(wb):
    """The raw monthly sheet — the only one whose header starts with FUND_ID.

    (The workbook also carries the author's pivot sheets, and the data sheet's
    name contains right-to-left marks, so it is found by content, not by name.)
    """
    for ws in wb.worksheets:
        first = next(ws.iter_rows(max_row=1, values_only=True), None)
        if first and norm(first[0]) == "FUND_ID":
            return ws
    raise RuntimeError("no data sheet with a FUND_ID header in " + wb.path)


def iter_data(fields):
    """Yield one dict per monthly row, with `year` / `month` already parsed.

    `fields` is the list of source columns to pull. Rows for excluded products
    are dropped here so no stage can forget the filter.
    """
    wb = openpyxl.load_workbook(DATA_XLSX, read_only=True, data_only=True)
    ws = _data_sheet(wb)
    it = ws.iter_rows(values_only=True)
    header = [norm(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    want = [(f, idx[f]) for f in fields]
    c_per, c_name = idx["REPORT_PERIOD"], idx["FUND_NAME"]
    for r in it:
        # REPORT_PERIOD is YYYYMM (string in this extract, datetime if re-saved).
        per = r[c_per]
        if per is None:
            continue
        if hasattr(per, "year"):
            year, month = per.year, per.month
        else:
            per = str(per).strip()
            if len(per) < 6 or not per[:6].isdigit():
                continue
            year, month = int(per[:4]), int(per[4:6])
        name = norm(r[c_name])
        if excluded_fund(name):
            continue
        row = {f: r[i] for f, i in want}
        row["fund_name"] = name
        row["year"] = year
        row["month"] = month
        yield row
    wb.close()


def load_risk_map():
    """FUND_NAME -> {specialization, sub_specialization, risk} from the map."""
    wb = openpyxl.load_workbook(RISKMAP_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = [norm(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in it:
        name = norm(r[idx["FUND_NAME"]])
        if not name:
            continue
        out[name] = {
            "specialization": norm(r[idx["SPECIALIZATION"]]),
            "sub_specialization": norm(r[idx["SUB_SPECIALIZATION"]]),
            "risk": norm_risk(r[idx["רמת סיכון"]]),
        }
    wb.close()
    return out


def num(v):
    """Float or None — the extract stores every number as text."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
