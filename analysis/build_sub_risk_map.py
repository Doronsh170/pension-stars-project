#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Map every `SUB_SPECIALIZATION` track to a single risk level.

`fund-risk-map.xlsx` labels one *fund* at a time, so it covers only the 704
funds trading when the snapshot was taken and says `ללא נתון` for 64 of them.
This stage lifts the label one level up, to the track: the 56 distinct
`SUB_SPECIALIZATION` values in the source cover all 1,273 funds in the
2010-2025 history, closed ones included, and a track never changes its label.

A track's level is decided from the evidence the source already carries:

  a. what the track's funds actually held -- the median equity share of assets
     (STOCK_MARKET_EXPOSURE / TOTAL_ASSETS) on the study's own cut-points, with
     the two no-equity levels split by median volatility;
  b. the level the curated per-fund map gives most of that track's funds.

(b) wins where the two agree or sit one level apart, so a track keeps the
label a human gave it; (a) wins where the per-fund map has no majority or
contradicts the holdings by two levels or more. Behaviour alone settles the
seven theme tracks the map leaves as `ללא נתון` (הלכה יהודית, חו"ל,
קיימות ...), and volatility alone settles the few tracks that never reported
an exposure.

Outputs, both sorted by risk level:
  analysis/sub_specialization_risk_map.csv
  sub-specialization-risk-map.xlsx   (mapping sheet + an evidence sheet)
"""
import csv
import os
import statistics
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from source import HERE, ROOT, RISK_ORDER, iter_data, load_risk_map, norm, num

OUT_CSV = os.path.join(HERE, "sub_specialization_risk_map.csv")
OUT_XLSX = os.path.join(ROOT, "sub-specialization-risk-map.xlsx")

# Equity share of assets, in percent -> risk level. Same cut-points the
# per-fund stage uses, so the two stages cannot drift apart.
EQ_CUTS = [(75, "גבוה"), (25, "בינוני"), (5, "נמוך-בינוני")]
SD_SPLIT_NO_EQUITY = 1.5   # below this, a no-equity track is `נמוך מאוד`
SD_CUTS = [(10, "גבוה"), (5.5, "בינוני"), (3.5, "נמוך-בינוני"), (1.5, "נמוך")]

# The per-fund map only speaks for a track when its labels mostly agree with
# each other, and only within one level of what the track's funds held.
MAP_MIN_PURITY = 0.60
MAP_MAX_DISTANCE = 1

# A track whose funds' equity shares span more than this many points is flagged:
# one level is a rough fit for it, whatever the rule picks.
WIDE_IQR = 20


def eq_band(eq, sd):
    """Risk level implied by the track's equity share (and, at 0%, its vol)."""
    for cut, level in EQ_CUTS:
        if eq >= cut:
            return level
    if sd is not None and sd < SD_SPLIT_NO_EQUITY:
        return "נמוך מאוד"
    return "נמוך"


def behaviour(eq, sd):
    """(risk level, source) implied by what the track's funds held."""
    if eq is not None:
        return eq_band(eq, sd), "חשיפה מנייתית"
    if sd is not None:
        for cut, level in SD_CUTS:
            if sd >= cut:
                return level, "תנודתיות"
        return "נמוך מאוד", "תנודתיות"
    return "", ""


def map_majority(map_levels):
    """The level most of the track's mapped funds carry, if most of them do."""
    n = sum(map_levels.values())
    if not n:
        return ""
    top, n_top = map_levels.most_common(1)[0]
    return top if n_top / n >= MAP_MIN_PURITY else ""


def decide(map_levels, eq, sd):
    """(risk level, source) for one track."""
    seen, src = behaviour(eq, sd)
    curated = map_majority(map_levels)
    if curated:
        far = seen and abs(RISK_ORDER.index(curated) - RISK_ORDER.index(seen))
        if not seen or far <= MAP_MAX_DISTANCE:
            return curated, "מיפוי קופות"
    return seen, src


def collect_funds():
    """FUND_ID -> the fund's track, its names, and its risk-proxy medians."""
    funds = defaultdict(lambda: {"names": {}, "spec": Counter(), "sub": Counter(),
                                 "eq": [], "sd": []})
    fields = ["FUND_ID", "SPECIALIZATION", "SUB_SPECIALIZATION",
              "STOCK_MARKET_EXPOSURE", "TOTAL_ASSETS", "STANDARD_DEVIATION"]
    for r in iter_data(fields):
        f = funds[str(r["FUND_ID"])]
        f["names"][(r["year"], r["month"])] = r["fund_name"]
        spec, sub = norm(r["SPECIALIZATION"]), norm(r["SUB_SPECIALIZATION"])
        if spec:
            f["spec"][spec] += 1
        if sub:
            f["sub"][sub] += 1
        assets, exposure = num(r["TOTAL_ASSETS"]), num(r["STOCK_MARKET_EXPOSURE"])
        if assets and exposure is not None:
            f["eq"].append(exposure / assets * 100)
        sd = num(r["STANDARD_DEVIATION"])
        if sd is not None:
            f["sd"].append(sd)
    return funds


def quartiles(values):
    """(p25, median, p75) of a list, or (None, None, None)."""
    if not values:
        return None, None, None
    v = sorted(values)

    def pct(p):
        i = (len(v) - 1) * p
        lo = int(i)
        hi = min(lo + 1, len(v) - 1)
        return v[lo] + (v[hi] - v[lo]) * (i - lo)

    return pct(0.25), pct(0.5), pct(0.75)


def build():
    """One row per SUB_SPECIALIZATION value, sorted by risk level."""
    risk_map = load_risk_map()
    tracks = defaultdict(lambda: {"funds": 0, "eq": [], "sd": [], "levels": Counter(),
                                  "no_data": 0, "specs": Counter()})
    for f in collect_funds().values():
        if not f["sub"]:
            continue                       # the two funds with no track at all
        sub = f["sub"].most_common(1)[0][0]
        t = tracks[sub]
        t["funds"] += 1
        t["specs"][f["spec"].most_common(1)[0][0] if f["spec"] else ""] += 1
        names = [f["names"][k] for k in sorted(f["names"], reverse=True)]
        mapped = next((risk_map[n] for n in names if n in risk_map), None)
        if mapped:
            if mapped["risk"] in RISK_ORDER:
                t["levels"][mapped["risk"]] += 1
            else:
                t["no_data"] += 1          # the map's `ללא נתון`
        if f["eq"]:
            t["eq"].append(statistics.median(f["eq"]))
        if f["sd"]:
            t["sd"].append(statistics.median(f["sd"]))

    rows = []
    for sub, t in tracks.items():
        p25, med, p75 = quartiles(t["eq"])
        sd_med = statistics.median(t["sd"]) if t["sd"] else None
        level, src = decide(t["levels"], med, sd_med)
        seen, _ = behaviour(med, sd_med)
        n_map = sum(t["levels"].values())
        top_n = t["levels"].most_common(1)[0][1] if n_map else 0
        rows.append({
            "sub_specialization": sub,
            "risk_level": level,
            "funds": t["funds"],
            "equity_pct_median": None if med is None else round(med, 1),
            "std_dev_median": None if sd_med is None else round(sd_med, 2),
            "source": src,
            "note": note(t, med, p25, p75, level, seen),
            # evidence, for the second sheet
            "specializations": " · ".join(
                f"{s or '—'} ({n})" for s, n in t["specs"].most_common()),
            "mapped_funds": n_map,
            "map_no_data": t["no_data"],
            "map_agreement": "" if not n_map else f"{round(top_n / n_map * 100)}%",
            "map_distribution": " · ".join(
                f"{lv} {n}" for lv, n in sorted(
                    t["levels"].items(), key=lambda kv: RISK_ORDER.index(kv[0]))),
            "equity_pct_p25": None if p25 is None else round(p25, 1),
            "equity_pct_p75": None if p75 is None else round(p75, 1),
        })

    rows.sort(key=lambda r: (RISK_ORDER.index(r["risk_level"]) if r["risk_level"]
                             in RISK_ORDER else len(RISK_ORDER),
                             -(r["equity_pct_median"] or 0), r["sub_specialization"]))
    return rows


def note(t, med, p25, p75, level, seen):
    """Why this track deserves a second look, if it does."""
    notes = []
    if p25 is not None and p75 - p25 > WIDE_IQR:
        specs = " · ".join(f"{s}" for s, _ in t["specs"].most_common(3))
        notes.append(f"מסלול לא אחיד: חשיפה מנייתית {p25:.0f}%–{p75:.0f}% "
                     f"(מופיע תחת {specs}) — רמה אחת היא קירוב")
    n_map = sum(t["levels"].values())
    if n_map:
        top = t["levels"].most_common(1)[0][0]
        if not map_majority(t["levels"]):
            notes.append("מיפוי הקופות אינו חד־משמעי — הרמה נקבעה לפי החשיפה בפועל")
        elif top != level:
            notes.append(f"מיפוי הקופות נטה ל־{top} ({n_map} מתוך {t['funds']} קופות), "
                         f"אך החשיפה בפועל ({med:.0f}%) רחוקה ממנה ביותר מרמה אחת")
        # Below 5% equity the exposure cannot tell the bottom levels apart, so
        # the map refining them is expected, not a divergence worth flagging.
        elif seen and seen != level and med is not None and med >= 5:
            notes.append(f"נשמרה הרמה שבמיפוי הקופות; החשיפה בפועל "
                         f"({med:.0f}%) מתאימה ל־{seen}")
    elif t["no_data"]:
        notes.append(f"{t['no_data']} קופות מסומנות \"ללא נתון\" במיפוי — "
                     "הרמה נגזרה מהתנהגות הקופות")
    return "; ".join(notes)


MAP_COLS = [
    ("sub_specialization", "SUB_SPECIALIZATION", 44),
    ("risk_level", "רמת סיכון", 13),
    ("funds", "מספר קופות", 11),
    ("equity_pct_median", "חציון חשיפה מנייתית %", 20),
    ("std_dev_median", "חציון סטיית תקן", 15),
    ("source", "מקור השיוך", 15),
    ("note", "הערה", 70),
]
EVIDENCE_COLS = [
    ("sub_specialization", "SUB_SPECIALIZATION", 44),
    ("risk_level", "רמת סיכון", 13),
    ("specializations", "SPECIALIZATION (מספר קופות)", 46),
    ("funds", "מספר קופות", 11),
    ("mapped_funds", "מתוכן ממופות בטבלת הקופות", 24),
    ("map_no_data", "מתוכן \"ללא נתון\"", 16),
    ("map_agreement", "אחידות המיפוי", 14),
    ("map_distribution", "פילוח רמות במיפוי הקופות", 34),
    ("equity_pct_p25", "חשיפה מנייתית — רבעון תחתון", 26),
    ("equity_pct_median", "חציון", 10),
    ("equity_pct_p75", "רבעון עליון", 13),
    ("std_dev_median", "חציון סטיית תקן", 15),
]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
BAND = {"נמוך מאוד": "E8F1FB", "נמוך": "E6F4EA", "נמוך-בינוני": "FFF6E0",
        "בינוני": "FDEDE1", "גבוה": "FCE4E4"}
THIN = Side(style="thin", color="D0D0D0")


def write_sheet(ws, rows, cols):
    ws.sheet_view.rightToLeft = True
    ws.append([label for _, label, _ in cols])
    for i, (_, _, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r[key] if r[key] is not None else "" for key, _, _ in cols])
        fill = BAND.get(r["risk_level"])
        for c in ws[ws.max_row]:
            c.border = Border(bottom=THIN)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        ws.cell(ws.max_row, 2).font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32


def main():
    rows = build()

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as fh:
        keys = [k for k, _, _ in MAP_COLS] + [
            k for k, _, _ in EVIDENCE_COLS if k not in {c[0] for c in MAP_COLS}]
        w = csv.DictWriter(fh, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    wb = openpyxl.Workbook()
    write_sheet(wb.active, rows, MAP_COLS)
    wb.active.title = "מיפוי מסלולים"
    write_sheet(wb.create_sheet("נתוני הבסיס"), rows, EVIDENCE_COLS)
    wb.save(OUT_XLSX)

    print(f"tracks: {len(rows)}   funds covered: {sum(r['funds'] for r in rows)}")
    by_level = Counter(r["risk_level"] for r in rows)
    for level in RISK_ORDER:
        n_funds = sum(r["funds"] for r in rows if r["risk_level"] == level)
        print(f"  {level:11s}: {by_level[level]:2d} tracks, {n_funds:4d} funds")
    src = Counter(r["source"] for r in rows)
    print("  source:", ", ".join(f"{k} {v}" for k, v in src.most_common()))
    print(f"  flagged for review: {sum(1 for r in rows if r['note'])}")
    print("\nwrote", OUT_CSV, "\n      ", OUT_XLSX)


if __name__ == "__main__":
    main()
